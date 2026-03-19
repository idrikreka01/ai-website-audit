"""
Tests for Excel rubric artifact generation and storage.

Verifies that a workbook with Questions and Output tabs is written to
{domain}__{session_id}/output.xlsx and that an excel_rubric_xlsx artifact
record is created.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from worker.excel_rubric import save_excel_rubric_workbook


def _config(artifacts_dir: str) -> SimpleNamespace:
    return SimpleNamespace(artifacts_dir=artifacts_dir)


class _RepoStub:
    def __init__(self, session_id):
        self.session_id = session_id
        self.created_artifacts = []
        self.logs = []

    def list_questions(self, page_type=None, category=None):
        return [
            {
                "id": uuid4(),
                "question_id": 1,
                "key": "Q1",
                "stage": "Awareness",
                "category": "Homepage",
                "page_type": "homepage",
                "tier": 1,
                "severity": 2,
                "question": "Is the homepage clear?",
                "bar_chart_category": "Homepage Clarity",
                "exact_fix": "Clarify homepage.",
                "allowed_evidence_types": ["screenshot", "visible_text", "features_json"],
                "pass_criteria": "Homepage is clear.",
                "fail_criteria": "Homepage is confusing.",
                "notes": "",
            }
        ]

    def get_session_by_id(self, session_id):
        return {
            "id": session_id,
            "url": "https://example.com/",
        }

    def get_audit_results_by_session_id(self, session_id_str):
        return [
            {
                "result_id": 1,
                "question_id": 1,
                "session_id": session_id_str,
                "result": "pass",
                "reason": "Looks good",
                "confidence_score": 9,
            }
        ]

    def get_pages_by_session_id(self, session_id):
        return [
            {
                "id": uuid4(),
                "session_id": session_id,
                "page_type": "homepage",
                "viewport": "desktop",
                "status": "ok",
                "load_timings": {},
                "low_confidence_reasons": [],
            }
        ]

    def get_artifacts_by_session_id(self, session_id):
        pages = self.get_pages_by_session_id(session_id)
        page_id = pages[0]["id"]
        return [
            {
                "id": uuid4(),
                "session_id": session_id,
                "page_id": page_id,
                "type": "screenshot",
                "storage_uri": f"example.com__{session_id}/homepage/desktop/screenshot.png",
            },
            {
                "id": uuid4(),
                "session_id": session_id,
                "page_id": page_id,
                "type": "visible_text",
                "storage_uri": f"example.com__{session_id}/homepage/desktop/visible_text.txt",
            },
            {
                "id": uuid4(),
                "session_id": session_id,
                "page_id": page_id,
                "type": "features_json",
                "storage_uri": f"example.com__{session_id}/homepage/desktop/features_json.json",
            },
        ]

    def create_artifact(
        self,
        *,
        session_id,
        page_id,
        artifact_type,
        storage_uri,
        size_bytes,
        retention_until=None,
        checksum=None,
    ):
        self.created_artifacts.append(
            {
                "session_id": session_id,
                "page_id": page_id,
                "artifact_type": artifact_type,
                "storage_uri": storage_uri,
                "size_bytes": size_bytes,
                "retention_until": retention_until,
                "checksum": checksum,
            }
        )
        return self.created_artifacts[-1]

    def create_log(self, *, session_id, level, event_type, message, details=None):
        self.logs.append(
            {
                "session_id": session_id,
                "level": level,
                "event_type": event_type,
                "message": message,
                "details": details or {},
            }
        )


def test_save_excel_rubric_workbook_writes_output_and_creates_artifact(monkeypatch):
    """Excel rubric workbook is written to output.xlsx and artifact is recorded."""
    session_id = uuid4()
    domain = "example.com"
    repo = _RepoStub(session_id)

    with tempfile.TemporaryDirectory(dir=os.getcwd()) as tmpdir:
        from worker import storage as storage_mod
        from worker import excel_rubric as excel_mod

        cfg = _config(tmpdir)
        monkeypatch.setattr(storage_mod, "get_config", lambda: cfg)
        monkeypatch.setattr(excel_mod, "get_storage_uri", storage_mod.get_storage_uri)

        result = save_excel_rubric_workbook(repo, session_id, domain)

        assert result is True
        assert repo.created_artifacts
        art = repo.created_artifacts[0]
        assert art["artifact_type"] == "excel_rubric_xlsx"
        assert art["page_id"] is None
        assert art["storage_uri"].startswith(f"{domain}__{session_id}")
        assert art["storage_uri"].endswith("output.xlsx")
        assert art["retention_until"] is None

        path = Path(tmpdir) / f"{domain}__{session_id}" / "output.xlsx"
        assert path.exists()

        wb = load_workbook(path)
        assert "Questions" in wb.sheetnames
        assert "Output" in wb.sheetnames

        questions_ws = wb["Questions"]
        output_ws = wb["Output"]

        questions_header = [cell.value for cell in next(questions_ws.iter_rows(max_row=1))]
        output_header = [cell.value for cell in next(output_ws.iter_rows(max_row=1))]

        assert questions_header == [
            "Category",
            "Questions",
            "AI",
            "Model",
            "Tier",
            "Severity",
            "Page",
            "Bar Chart Category (In Audit)",
            "Exact Fix:",
        ]
        assert output_header == ["Category", "Questions", "AI grade"]

